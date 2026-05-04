import re
import logging
from datetime import datetime
import docx
import pytest
import pandas as pd
import openpyxl
from pages.auth_page import AuthPage
from pages.product_page import ProductPage
from pages.payment_page import PaymentPage
from utils.ui_actions import UIActions
from utils.word_generator import WordGenerator
from faker import Faker
import random
fake = Faker()

output_data_path = './testdata/testdata_input.xlsx'
sheet_name = 'end_to_end'
logging = logging.getLogger(__name__)


def read_test_data():
    df = pd.read_excel(output_data_path, sheet_name, dtype=str, keep_default_na=False)
    df = df.dropna(how="all")
    return df.to_dict(orient='records')

@pytest.mark.parametrize("test_data", read_test_data(), ids=lambda x: x.get("testcase_id", "unknown"))
def test_end_to_end(page, test_data):
    document = ""

    testcase = test_data["testcase_id"]
    row = re.sub('[^0-9]', '', testcase)
    row = int(row)

    WordGenerator.clear_temp_screenshot(path="screenshots/")
    WordGenerator.clear_temp_words(path="reports/")

    try:
        #=== Test Data ===#
        name = test_data["name"]
        email = test_data["email"]
        password = test_data["password"]
        product_name = test_data["product_name"]
        comment = test_data["comment"]
        cc_number = test_data["cc_number"]
        cc_name = test_data["cc_name"]
        cc_cvv = test_data["cc_cvv"]
        cc_exp_month = test_data["cc_exp_month"]
        cc_exp_year = test_data["cc_exp_year"]


        #=== Steps ===#
        document = WordGenerator.start_document(page, test_data)
        auth_page = AuthPage(page)
        # auth_page.login(name, email, password, document)
        auth_page.login_alt(name, email, password, document)

        product_page = ProductPage(page)
        product_page.buy_product_by_search(product_name, comment, document)

        payment_page = PaymentPage(page)
        payment_page.pay_and_confirm(document, cc_name, cc_number, cc_cvv, cc_exp_month, cc_exp_year)

        auth_page.log_out(document)


        # ===[ Write to Excel ]===
        # workbook = openpyxl.load_workbook(output_data_path)
        # sheet = workbook[sheet_name]

        # sheet.cell(row=row + 1, column=1).value = f"TC00{row}"
        # sheet.cell(row=row + 1, column=3).value = name
        # logging.info(f"Writing name: {name} to Excel at row {row + 1}, column 3")
        # sheet.cell(row=row + 1, column=4).value = email
        # logging.info(f"Writing email to Excel at row {row + 1}, column 4")
        # sheet.cell(row=row + 1, column=7).value = dob
        # sheet.cell(row=row + 1, column=10).value = firstname
        # sheet.cell(row=row + 1, column=11).value = lastname
        # sheet.cell(row=row + 1, column=12).value = company
        # sheet.cell(row=row + 1, column=13).value = address
        # sheet.cell(row=row + 1, column=15).value = state
        # sheet.cell(row=row + 1, column=16).value = zipcode
        # sheet.cell(row=row + 1, column=17).value = mobile

        # workbook.save(output_data_path)
        # logging.info(f"Successfully write data to Excel.")

        remarks = "PASS"
    except Exception as e:
        remarks = "FAIL"
        logging.error(e)
        raise e

    finally:
        now = datetime.now()
        now = now.strftime("%d-%m-%Y_%H-%M-%S")
        doc_report_name = f"reports/[REPORT] [{testcase}] END TO END - {now} - {remarks}.docx"
        document.save(doc_report_name)
        doc = docx.Document(doc_report_name)
        doc.save(doc_report_name)
        logging.info(f"Successfully generated report: {doc_report_name}")