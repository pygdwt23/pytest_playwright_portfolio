import re
import logging
from datetime import datetime
import docx
import pytest
import pandas as pd
import openpyxl
from pages.auth_page import AuthPage
from utils.ui_actions import UIActions
from utils.word_generator import WordGenerator
from utils.send_email import EmailSender
from faker import Faker
import random
from pathlib import Path

fake = Faker()

output_data_path = './testdata/testdata_input.xlsx'
sheet_name = 'registration'
PROJECT_ROOT = Path(__file__).parent.parent
DOWNLOAD_DIR = PROJECT_ROOT / "downloads"
ALLURE_DIR = PROJECT_ROOT / "allure-reports"
logging = logging.getLogger(__name__)


def read_test_data():
    df = pd.read_excel(output_data_path, sheet_name, dtype=str, keep_default_na=False)
    df = df.dropna(how="all")
    return df.to_dict(orient='records')

@pytest.mark.parametrize("test_data", read_test_data(), ids=lambda x: x.get("testcase_id", "unknown"))
def test_delete_account(page, test_data):
    document = ""

    testcase = test_data["testcase_id"]
    row = re.sub('[^0-9]', '', testcase)
    row = int(row)

    WordGenerator.clear_temp_screenshot(path="screenshots/")
    WordGenerator.clear_temp_words(path="reports/")
    WordGenerator.clear_downloaded_files(path=DOWNLOAD_DIR)
    WordGenerator.clear_downloaded_files(path=ALLURE_DIR)

    try:
        #=== Test Data ===#
        name = test_data["name"]
        email = test_data["email"]
        password = test_data["password"]
        title = test_data["title"]
        dob = test_data["dob"]
        newsletter = test_data["newsletter"]
        offers = test_data["offers"]
        firstname = test_data["firstname"]
        lastname = test_data["lastname"]
        company = test_data["company"]
        address = test_data["address"]
        country = test_data["country"]
        state = test_data["state"]
        zipcode = test_data["zipcode"]
        mobile = test_data["mobile"]


        #=== Data Generation ===#
        if name == "":
            if title == "Mr":
                name = fake.last_name_male()
            else:
                name = fake.last_name_female()

        dummy_email = fake.email()

        random_num = random.randint(100, 999)

        if password == "":
            password = fake.password(length=10, special_chars=True, digits=True, upper_case=True, lower_case=True)
        if dob == "":
            dob = fake.date(pattern="%d/%m/%Y")
            dob = datetime.strptime(dob, "%d/%m/%Y")
        birth_date = dob.strftime("%d")
        birth_month = dob.strftime("%B")
        birth_year = dob.strftime("%Y")
        logging.info(f"Date of birth: {birth_date} / {birth_month} / {birth_year}")

        if firstname == "":
            if title == "Mr":
                firstname = fake.first_name_male()
            else:
                firstname = fake.first_name_female()
        if lastname == "":
            lastname = name

        if email == "":
            email = f"{firstname}{lastname}{random_num}@tester.com"

        if company == "":
            company = fake.company()

        if address == "":
            address = fake.address()

        if country == "":
            country = ["India", "Australia", "Canada", "New Zealand", "Singapore"]
            country = fake.random_element(elements=country)

        if state == "":
            state = fake.state()

        if zipcode == "":
            zipcode = fake.postcode()

        if mobile == "":
            mobile = fake.phone_number()

        #=== Steps ===#
        document = WordGenerator.start_document(page, test_data)
        auth_page = AuthPage(page)
        auth_page.login_alt(name, email, password, document)
        auth_page.delete_account(document)
        # auth_page.signup(name, email, document)
        # firstname, lastname, company, address, state, zipcode, mobile = auth_page.app_form(password, title, birth_date, birth_month, birth_year, newsletter, offers, firstname, lastname, company, address, country, state, zipcode, mobile, email, document)


        # ===[ Write to Excel ]===
        # workbook = openpyxl.load_workbook(output_data_path)
        # sheet = workbook[sheet_name]

        # sheet.cell(row=row + 1, column=1).value = f"TC00{row}"
        # sheet.cell(row=row + 1, column=3).value = name
        # logging.info(f"Writing name: {name} to Excel at row {row + 1}, column 3")
        # sheet.cell(row=row + 1, column=4).value = email
        # logging.info(f"Writing email to Excel at row {row + 1}, column 4")
        # sheet.cell(row=row + 1, column=7).value = dob
        # sheet.cell(row=row + 1, column=10).value = firstname
        # sheet.cell(row=row + 1, column=11).value = lastname
        # sheet.cell(row=row + 1, column=12).value = company
        # sheet.cell(row=row + 1, column=13).value = address
        # sheet.cell(row=row + 1, column=15).value = state
        # sheet.cell(row=row + 1, column=16).value = zipcode
        # sheet.cell(row=row + 1, column=17).value = mobile

        # workbook.save(output_data_path)
        # logging.info(f"Successfully write data to Excel.")

        remarks = "PASS"
    except Exception as e:
        remarks = "FAIL"
        logging.error(e)
        raise e

    finally:
        now = datetime.now()
        now = now.strftime("%d-%m-%Y_%H-%M-%S")
        doc_report_name = f"reports/[REPORT] DELETE ACCOUNT - {now} - {remarks}.docx"
        document.save(doc_report_name)
        doc = docx.Document(doc_report_name)
        doc.save(doc_report_name)
        logging.info(f"Successfully generated report: {doc_report_name}")

        send_email = EmailSender()
        send_email.send_report(
            to="pygdwt@gmail.com",
            test_name=f"DELETE ACCOUNT - {testcase}",
            status=remarks,
            report_path=doc_report_name,
            extra_body=f"Dear recipient,\n\nPlease find attached the test report for your recent test execution.\n\nBest regards,\nQA Team"
        )